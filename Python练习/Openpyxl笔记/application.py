import openpyxl

def readwb(wbname, sheetname=None):
    # 打开工作薄
    wb = openpyxl.load_workbook(wbname)
    # 获取要操作的工作表
    if not sheetname:
        sheet = wb.active
    else:
        sheet = wb[sheetname]

    # 获取商品信息保存到列表中
    #[ ['name', price, count] ]
    all_info = []
    for row in sheet.rows:
        child = [cell.value for cell in row]
        all_info.append(child)
    return sorted(all_info, key=lambda item: item[1])


def save_to_excel(data, wbname, sheetname='sheet1'):
    """
    将信息保存到excel表中;
    [[' BOOK', 50, 3], ['APPLE', 100, 1], ['BANANA', 200, 0.5]]
    """
    print("写入Excel[%s]中......." %(wbname))
    # 打开excel表, 如果文件不存在， 自己实例化一个WorkBook对象
    wb = openpyxl.Workbook()
    # 修改当前工作表的名称
    sheet = wb.active
    # 修改工作表的名称
    sheet.title = sheetname

    for row, item in enumerate(data):  # 0 [' BOOK', 50, 3]
        for column, cellValue in enumerate(item): #  0 ' BOOK'
            sheet.cell(row=row+1, column=column+1, value=cellValue)

    # ** 往单元格写入内容
    # sheet.cell['B1'].value = "value"
    # sheet.cell(row=1, column=2, value="value")
    # 保存写入的信息
    wb.save(filename=wbname)
    print("写入成功!")

data = readwb(wbname='Book1.xlsx')
save_to_excel(data, wbname='Book2.xlsx', sheetname="排序商品信息")